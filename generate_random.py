import openpyxl as xl
import random

def generate(total):
    wb=xl.Workbook()
    ws=wb['Sheet']
    ws.append(['Name','Age','M1','M2','M3'])

    for i in range(2,total+2):
        cell=ws.cell(i,1)
        cell.value='name'+str(i-1)

    for i in range(2,total+2):
        cell=ws.cell(i,2)
        cell.value=random.randint(18,30)

    for i in range(2,total+2):
        for j in range(3,6):
            cell=ws.cell(i,j)
            cell.value=random.randint(30,100)

    name=input('Enter the name with which you want to save the file :')
    wb.save('./excelsheets/'+name+'.xlsx')

try:
    total=eval(input('Enter the number of records you want :'))
    if(total>0):
        generate(total)
    else:
        print('Please enter a valid integer')
except ValueError:
    print('Please enter a valid integer')



