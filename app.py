import openpyxl as xl
import os
from pathlib import Path
from openpyxl.chart import BarChart,Reference

#listing all excel files inside the folder
print('Excelsheets present with us : ')
path=Path('./excelsheets/')
list=path.glob('*.xlsx')
for file in list:
    print(os.path.basename(file))

#get the file which needs to be modified
filename=str(input('Enter the complete filename you want to process :'))
wb=xl.load_workbook('./excelsheets/'+filename)

print("Please enter '3' for modifying marks in subject3, '4' for marks in subject4, '5' for marks in subject5")


def modify(choice):
    sheet = wb['Sheet']
    # we can get a cell as sheet['a1'] or use sheet.cell with rownum,colnum

    rows = sheet.max_row
    cols = sheet.max_column

    # printing sheet data

    print('Current sheet data :')
    for i in range(1, rows + 1):
        for j in range(1, cols + 1):
            print(sheet.cell(i, j).value, end="  ")
        print()


    # accessing marks1 and changing them to x% of original value
    try:
        print('Tell the percentage to which you want the current marks to get changed : ',end=' ')
        percent=float(input())
        new_cell = sheet.cell(1, 6)
        new_cell.value = 'New M'+str(choice-2)

        for i in range(2, rows + 1):
            cell = sheet.cell(i, choice)
            if(cell.value is not None):
                correct_value = cell.value * (percent/100)
                corrected_value_cell = sheet.cell(i, 6)
                corrected_value_cell.value = correct_value

        values1 = Reference(sheet, min_row=2, max_row=rows, min_col=choice, max_col=choice)
        gsheet=wb.create_sheet('graph_sheet')
        chart1 = BarChart()
        chart1.title='Bar chart for previous values of M'+str(choice-2)
        chart1.y_axis.title='marks'
        chart1.x_axis.title='students'
        chart1.add_data(values1)
        gsheet.add_chart(chart1, 'C5')

        values2 = Reference(sheet, min_row=2, max_row=rows, min_col=6, max_col=6)
        chart2 = BarChart()
        chart2.title = 'Bar chart for new values of M'+str(choice-2)
        chart2.y_axis.title = 'marks'
        chart2.x_axis.title = 'students'
        chart2.add_data(values2)
        gsheet.add_chart(chart2, 'L5')
        # saving the workbook
        wb.save('./modified_sheets/' + filename)

    except ValueError:
        print('Insert percentage only as numerals')

try:
    choice=int(input())
    if(choice==3 or choice ==4 or choice==5):
        modify(choice)
    else:
        print('You can choose to modify only column 3 or 4 or 5')

except ValueError:
    print('Please enter a valid integer')
